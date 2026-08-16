import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
BODY_FONT = Font(size=11)
THIN_BORDER = None


def _style_sheet(ws, df):
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col_idx, col in enumerate(df.columns, start=1):
        if df.empty:
            max_len = len(str(col))
        else:
            max_len = max((len(str(x)) for x in df[col].tolist()), default=len(str(col)))
            max_len = max(max_len, len(str(col)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 4, 12), 45)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="top")


def save_dataframes(dfs: list, path: str):
    wb = Workbook()
    wb.remove(wb.active)
    for name, df in dfs:
        ws = wb.create_sheet(title=str(name)[:31])
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        _style_sheet(ws, df)
    wb.save(path)
    return path


def read_dataframes(path: str) -> list:
    sheets = pd.read_excel(path, sheet_name=None)
    return [(name, df) for name, df in sheets.items()]
